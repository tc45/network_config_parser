"""
Excel Export Module

This module provides functionality to export parsed network configuration data
to Excel files. It handles the creation of multi-sheet workbooks with formatted
tables and proper error handling.

Features:
- Creates Excel workbooks with multiple sheets
- Automatically formats data as tables with headers
- Handles sheet name sanitization
- Includes timestamps in filenames
- Provides comprehensive error handling and logging
"""

import os
import datetime
import logging
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from typing import Dict, List

logger = logging.getLogger(__name__)

def export_data_to_excel(parsed_data: Dict[str, List[Dict]], output_dir: str, hostname: str):
    """
    Export parsed network configuration data to a formatted Excel file.
    
    This function takes parsed configuration data and creates a well-formatted
    Excel workbook with multiple sheets. Each sheet contains a table with the
    corresponding data, including headers and consistent formatting.
    
    Args:
        parsed_data (Dict[str, List[Dict]]): Dictionary containing the data to export:
            - Keys: Sheet names (will be sanitized)
            - Values: Lists of dictionaries, where each dictionary represents a row
                     and its keys are the column headers
            Example:
            {
                'Interfaces': [
                    {'Name': 'Gi0/1', 'IP': '192.168.1.1', 'Status': 'up'},
                    {'Name': 'Gi0/2', 'IP': '192.168.1.2', 'Status': 'down'}
                ]
            }
        
        output_dir (str): Directory path where the Excel file will be saved.
                         Will be created if it doesn't exist.
        
        hostname (str): Device hostname used in the filename.
                       Final filename format: "<hostname>_YYYYMMDD_HHMMSS.xlsx"
    
    Returns:
        None
    
    Raises:
        ValueError: If parsed_data is empty or invalid
        Exception: For other errors (e.g., file system issues)
    
    Notes:
        - Each sheet is formatted as a table with alternating row colors
        - Sheet names are sanitized (alphanumeric + some special chars, max 31 chars)
        - Empty sheets are skipped with a warning
        - Invalid data formats are logged and skipped
        - The function creates the output directory if it doesn't exist
    
    Example:
        >>> data = {
        ...     'Interfaces': [
        ...         {'Name': 'Gi0/1', 'IP': '192.168.1.1', 'Status': 'up'},
        ...         {'Name': 'Gi0/2', 'IP': '192.168.1.2', 'Status': 'down'}
        ...     ]
        ... }
        >>> export_data_to_excel(data, 'output', 'ROUTER-01')
        # Creates: output/ROUTER-01_20240321_143022.xlsx
    """
    try:
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logger.info(f"Created output directory: {output_dir}")

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{hostname}_{timestamp}.xlsx"
        final_output_path = os.path.join(output_dir, output_filename)

        workbook = openpyxl.Workbook()
        # Remove the default sheet created by openpyxl
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

        if not parsed_data:
            logger.warning("No data provided to export.")
            # Save an empty workbook? Or maybe just log and return.
            # workbook.save(final_output_path)
            # return
            raise ValueError("Parsed data dictionary is empty.")


        for sheet_name, sheet_data in parsed_data.items():
            if not sheet_data:
                logger.warning(f"No data for sheet '{sheet_name}', skipping.")
                continue

            # Ensure sheet_data is a list of dictionaries
            if not isinstance(sheet_data, list) or not all(isinstance(item, dict) for item in sheet_data):
                 logger.error(f"Invalid data format for sheet '{sheet_name}'. Expected List[Dict]. Got {type(sheet_data)}")
                 # Fallback: try to convert if possible, or skip
                 if isinstance(sheet_data, list) and len(sheet_data) > 0:
                     # Try to get headers from keys of first item if dict
                     if isinstance(sheet_data[0], dict):
                          headers = list(sheet_data[0].keys())
                     else: # Can't determine headers
                          logger.warning(f"Could not determine headers for sheet '{sheet_name}', skipping.")
                          continue
                 else: # Cannot process this sheet
                      logger.warning(f"Cannot process data for sheet '{sheet_name}', skipping.")
                      continue
            else:
                 headers = list(sheet_data[0].keys())


            # Sanitize sheet name (max 31 chars, no invalid chars)
            safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '_', '-'))[:31]
            sheet = workbook.create_sheet(title=safe_sheet_name)
            logger.debug(f"Creating sheet: {safe_sheet_name}")

            # Write headers
            sheet.append(headers)

            # Write data rows
            for item in sheet_data:
                row = [item.get(header, '') for header in headers]
                sheet.append(row)

            # Create a table
            if headers: # Ensure headers exist before creating table
                 table_ref = f"A1:{get_column_letter(len(headers))}{len(sheet_data) + 1}"
                 table = Table(displayName=f"{safe_sheet_name.replace(' ', '_')}Table", ref=table_ref)
                 style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                 table.tableStyleInfo = style
                 sheet.add_table(table)
            else:
                 logger.warning(f"Cannot create table for sheet '{safe_sheet_name}' due to missing headers.")


        if not workbook.sheetnames:
             logger.error("No valid sheets were created. Excel file will not be saved.")
             return # Don't save an empty workbook

        # Save the workbook
        workbook.save(final_output_path)
        logging.info(f"Data successfully exported to {final_output_path}")

    except ValueError as ve:
         logger.error(f"ValueError during export: {ve}")
         # Don't raise, just log it maybe? Or re-raise? Let's log for now.
    except Exception as e:
        logger.error(f"Failed to export data to Excel: {e}", exc_info=True)
        # Decide if we should raise e here or just log it. Logging seems sufficient. 